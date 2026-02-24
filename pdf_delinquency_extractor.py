"""
통일경영공시 PDF에서 은행별 연체대출비율을 추출하여 엑셀로 정리하는 모듈.
- 자산건전성 지표 항목 내 연체대출비율(금액기준)을 타겟으로 추출
- Gemini OCR 기반 추출 (1차)
- pdfplumber 기반 테이블/텍스트 추출 (fallback, 자산건전성 지표 페이지 우선)
- 별도 엑셀 파일로 생성
"""

import os
import re
import json
import time
import logging
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    from google import genai
    from google.genai import types
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

logger = logging.getLogger(__name__)

# 자산건전성 지표 섹션 감지 키워드
_ASSET_QUALITY_KEYWORDS = [
    "자산건전성지표",
    "자산건전성 지표",
    "자산건전성",
]

# 금액기준 연체대출비율 우선 키워드
_AMOUNT_BASED_KEYWORDS = [
    "연체대출비율(금액기준)",
    "연체대출채권비율(금액기준)",
    "연체대출금비율(금액기준)",
    "연체율(금액기준)",
    "연체비율(금액기준)",
]

# 일반 연체대출비율 키워드 (자산건전성 지표 내에서 탐색)
_GENERIC_DELINQUENCY_KEYWORDS = [
    "연체대출비율",
    "연체대출채권비율",
    "연체대출금비율",
    "연체율",
    "연체비율",
]

# 건수기준 키워드 (제외 대상)
_COUNT_BASED_KEYWORDS = [
    "건수기준",
    "건기준",
]


def _classify_delinquency_cell(text: str) -> Optional[str]:
    """
    셀 텍스트가 연체율 관련 항목인지 판별하고, 유형을 반환한다.

    Returns:
        "amount" (금액기준 — 우선), "generic" (일반), "count" (건수기준 — 후순위), None (무관)
    """
    if not text:
        return None
    cleaned = text.strip().replace(" ", "").replace("\n", "")

    # 1. 금액기준 명시 → 최우선
    if any(kw.replace(" ", "") in cleaned for kw in _AMOUNT_BASED_KEYWORDS):
        return "amount"

    # 2. 건수기준 명시 → 후순위 (가능하면 제외)
    if any(kw in cleaned for kw in _COUNT_BASED_KEYWORDS):
        return "count"

    # 3. 일반 연체율 키워드
    if any(kw in cleaned for kw in _GENERIC_DELINQUENCY_KEYWORDS):
        return "generic"

    return None


# ============================================================
# Gemini OCR 기반 연체율 추출
# ============================================================

def _extract_with_gemini(pdf_path: str, api_key: str, log_callback=None) -> Optional[Dict[str, str]]:
    """
    Gemini API의 OCR 기능으로 PDF에서 금액기준 연체율을 추출한다.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    try:
        client = genai.Client(api_key=api_key)

        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()

        if len(pdf_bytes) > 20 * 1024 * 1024:
            log("    [Gemini] PDF 크기가 20MB를 초과하여 건너뜁니다.")
            return None

        prompt = (
            "이 통일경영공시 PDF에서 '자산건전성 지표' 항목 내 '연체대출비율'을 찾아주세요.\n\n"
            "중요 사항:\n"
            "1. 반드시 '자산건전성 지표' 섹션에 있는 '연체대출비율'만 추출하세요.\n"
            "2. PDF에 '건수기준'과 '금액기준' 두 종류가 있을 수 있습니다.\n"
            "   반드시 **금액기준** 연체대출비율만 추출하세요. 건수기준은 절대 사용하지 마세요.\n"
            "3. '연체대출비율', '연체대출채권비율', '연체율' 등 표현이 다를 수 있으나 "
            "모두 같은 항목입니다.\n\n"
            "공시기준(당기) 값과 전년동기(전기) 값을 각각 추출해주세요.\n"
            "반드시 아래 JSON 형식으로만 응답하세요:\n"
            '{"연체율_당기": "숫자", "연체율_전기": "숫자"}\n'
            "찾을 수 없으면 null로 표시하세요.\n"
            "숫자는 퍼센트(%) 단위의 소수점 숫자만 넣으세요. (예: 2.35)"
        )

        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[
                types.Part.from_bytes(
                    data=pdf_bytes,
                    mime_type="application/pdf"
                ),
                prompt
            ],
            config=types.GenerateContentConfig(
                temperature=0.1,
                max_output_tokens=256,
                response_mime_type="application/json",
            ),
        )

        result_text = response.text.strip()
        if "```json" in result_text:
            result_text = result_text.split("```json")[1].split("```")[0].strip()
        elif "```" in result_text:
            result_text = result_text.split("```")[1].split("```")[0].strip()

        data = json.loads(result_text)

        # 유효성 검증
        result = {}
        for key in ["연체율_당기", "연체율_전기"]:
            val = data.get(key)
            if val is not None and val != "null" and str(val).strip():
                try:
                    num = float(str(val).replace("%", "").strip())
                    if 0 <= num <= 100:
                        result[key] = str(num)
                except (ValueError, TypeError):
                    pass

        if result:
            log(f"    [Gemini OCR] 금액기준 연체율 추출 성공: 당기 {result.get('연체율_당기', '-')}% / 전기 {result.get('연체율_전기', '-')}%")
            return result
        else:
            log(f"    [Gemini OCR] 유효한 연체율 값 없음")
            return None

    except Exception as e:
        log(f"    [Gemini OCR] 오류: {e}")
        return None


# ============================================================
# pdfplumber 기반 연체율 추출 (fallback)
# ============================================================

def extract_delinquency_from_pdf(
    pdf_path: str,
    api_key: str = None,
    log_callback=None
) -> Optional[Dict[str, str]]:
    """
    단일 통일경영공시 PDF에서 연체율 값을 추출한다.
    1차: Gemini OCR (api_key가 있을 때)
    2차: pdfplumber 테이블 추출 (금액기준 우선)
    3차: pdfplumber 텍스트 정규식
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    if not os.path.exists(pdf_path):
        return None

    # 1차: Gemini OCR 시도
    if api_key and GEMINI_AVAILABLE:
        result = _extract_with_gemini(pdf_path, api_key, log_callback)
        if result:
            return result

    # 2차/3차: pdfplumber fallback
    if not PDFPLUMBER_AVAILABLE:
        log("    pdfplumber가 설치되지 않아 fallback 추출을 건너뜁니다.")
        return None

    try:
        with pdfplumber.open(pdf_path) as pdf:
            # 자산건전성 지표 섹션이 있는 페이지를 우선 탐색
            asset_quality_pages = []
            other_pages = []
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                text_clean = text.replace(" ", "")
                if any(kw.replace(" ", "") in text_clean for kw in _ASSET_QUALITY_KEYWORDS):
                    asset_quality_pages.append((page_num, page))
                else:
                    other_pages.append((page_num, page))

            # 2차: 자산건전성 지표 페이지 테이블에서 우선 탐색
            for page_num, page in asset_quality_pages:
                result = _search_delinquency_in_page(page)
                if result:
                    log(f"    [pdfplumber 테이블] 자산건전성 지표 페이지 {page_num + 1}에서 연체대출비율 발견")
                    return result

            # 2-1차: 자산건전성 페이지가 없으면 전체 페이지 테이블 탐색
            for page_num, page in other_pages:
                result = _search_delinquency_in_page(page)
                if result:
                    log(f"    [pdfplumber 테이블] 페이지 {page_num + 1}에서 연체대출비율 발견")
                    return result

            # 3차: 자산건전성 지표 페이지 텍스트 탐색
            for page_num, page in asset_quality_pages:
                result = _search_delinquency_in_text(page)
                if result:
                    log(f"    [pdfplumber 텍스트] 자산건전성 지표 페이지 {page_num + 1}에서 연체대출비율 발견")
                    return result

            # 3-1차: 전체 페이지 텍스트 탐색
            for page_num, page in other_pages:
                result = _search_delinquency_in_text(page)
                if result:
                    log(f"    [pdfplumber 텍스트] 페이지 {page_num + 1}에서 연체대출비율 발견")
                    return result

            # 실패 시 디버그 정보
            log(f"    [디버그] 총 {len(pdf.pages)}페이지 검색했으나 연체대출비율 미발견 (자산건전성 지표 페이지: {len(asset_quality_pages)}개)")

    except Exception as e:
        logger.error(f"PDF 파싱 오류 ({pdf_path}): {e}")
        log(f"    PDF 파싱 오류: {e}")

    return None


def _search_delinquency_in_page(page) -> Optional[Dict[str, str]]:
    """
    페이지의 테이블에서 연체율 행을 찾는다.
    우선순위: 금액기준 > 일반 > 건수기준 (건수기준은 금액기준/일반이 없을 때만)
    """
    tables = page.extract_tables()
    if not tables:
        return None

    # 모든 테이블에서 연체율 후보를 수집
    candidates = []  # [(priority, result), ...]

    for table in tables:
        if not table:
            continue

        # 헤더 행 감지
        header_row_idx = 0
        for h_idx in range(min(3, len(table))):
            row = table[h_idx]
            if row and any(
                cell and any(kw in cell.strip().replace(" ", "").replace("\n", "")
                             for kw in ["전기", "전년", "당기", "금기", "공시기준", "전년동기"])
                for cell in row if cell
            ):
                header_row_idx = h_idx
                break

        header_row = table[header_row_idx] if header_row_idx < len(table) else table[0]

        for row_idx, row in enumerate(table):
            if not row or row_idx == header_row_idx:
                continue
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue
                cell_type = _classify_delinquency_cell(cell)
                if cell_type is None:
                    continue

                result = _extract_values_from_row(table, row_idx, row, col_idx, header_row)
                if result:
                    # 우선순위: amount=0 (최우선), generic=1, count=2 (최후순위)
                    priority = {"amount": 0, "generic": 1, "count": 2}[cell_type]
                    candidates.append((priority, result))

    if not candidates:
        return None

    # 최우선순위 결과 반환
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]


def _extract_values_from_row(
    table: list, row_idx: int, row: list, label_col: int,
    header_row: list = None
) -> Optional[Dict[str, str]]:
    """연체율이 발견된 행에서 수치 값을 추출한다."""
    numbers = []
    for col_idx, cell in enumerate(row):
        if col_idx == label_col:
            continue
        if cell:
            val = _parse_number(cell)
            if val is not None:
                numbers.append((col_idx, val))

    if not numbers:
        if row_idx + 1 < len(table):
            next_row = table[row_idx + 1]
            if next_row:
                for col_idx, cell in enumerate(next_row):
                    if cell:
                        val = _parse_number(cell)
                        if val is not None:
                            numbers.append((col_idx, val))

    if not numbers:
        return None

    if header_row is None:
        header_row = table[0] if table else []
    prior_cols, current_cols = _identify_period_columns(header_row)

    result = {}
    if prior_cols and current_cols:
        for col_idx, val in numbers:
            if col_idx in current_cols:
                result["연체율_당기"] = str(val)
            elif col_idx in prior_cols:
                result["연체율_전기"] = str(val)
    elif len(numbers) >= 2:
        numbers.sort(key=lambda x: x[0])
        result["연체율_당기"] = str(numbers[0][1])
        result["연체율_전기"] = str(numbers[1][1])
    elif len(numbers) == 1:
        result["연체율_당기"] = str(numbers[0][1])

    return result if result else None


def _identify_period_columns(header_row: list) -> Tuple[set, set]:
    """헤더 행에서 전기/당기 컬럼 인덱스를 판별한다."""
    prior_cols = set()
    current_cols = set()

    if not header_row:
        return prior_cols, current_cols

    for idx, cell in enumerate(header_row):
        if not cell:
            continue
        cell_clean = cell.strip().replace(" ", "").replace("\n", "")
        if any(kw in cell_clean for kw in ["전기", "전년", "전분기", "전년동기"]):
            prior_cols.add(idx)
        elif any(kw in cell_clean for kw in ["당기", "당분기", "금기", "금분기", "공시기준"]):
            current_cols.add(idx)

    return prior_cols, current_cols


def _search_delinquency_in_text(page) -> Optional[Dict[str, str]]:
    """페이지 텍스트에서 연체율 값을 정규식으로 추출한다. 금액기준 우선."""
    text = page.extract_text()
    if not text:
        return None

    text_clean = text.replace(" ", "")

    # 1순위: 금액기준 명시적 패턴
    for kw in ["연체대출비율(금액기준)", "연체대출채권비율(금액기준)", "연체대출금비율(금액기준)", "연체율(금액기준)"]:
        kw_clean = kw.replace(" ", "")
        if kw_clean not in text_clean:
            continue
        pattern = re.escape(kw_clean) + r"[^\d]*?([\d]+(?:\.[\d]+)?)[%\s]+([\d]+(?:\.[\d]+)?)"
        match = re.search(pattern, text_clean)
        if match:
            return {"연체율_당기": match.group(1), "연체율_전기": match.group(2)}
        pattern_one = re.escape(kw_clean) + r"[^\d]*?([\d]+(?:\.[\d]+)?)"
        match = re.search(pattern_one, text_clean)
        if match:
            return {"연체율_당기": match.group(1)}

    # 2순위: 일반 연체율 키워드 (건수기준이 아닌 경우)
    # "건수기준"이 근처에 없는 연체율 패턴 매칭
    for kw in _GENERIC_DELINQUENCY_KEYWORDS:
        if kw not in text_clean:
            continue

        # 키워드 위치 찾기
        pos = text_clean.find(kw)
        if pos == -1:
            continue

        # 주변에 "건수기준"이 있으면 건너뛰기
        context = text_clean[max(0, pos - 15):pos + len(kw) + 15]
        if any(ck in context for ck in _COUNT_BASED_KEYWORDS):
            continue

        pattern_two = re.escape(kw) + r"[^\d]*?([\d]+(?:\.[\d]+)?)[%\s]+([\d]+(?:\.[\d]+)?)"
        match = re.search(pattern_two, text_clean[pos:])
        if match:
            return {"연체율_당기": match.group(1), "연체율_전기": match.group(2)}

        pattern_one = re.escape(kw) + r"[^\d]*?([\d]+(?:\.[\d]+)?)"
        match = re.search(pattern_one, text_clean[pos:])
        if match:
            return {"연체율_당기": match.group(1)}

    return None


def _parse_number(text: str) -> Optional[float]:
    """텍스트에서 숫자(연체율 %) 값을 파싱한다."""
    if not text:
        return None
    cleaned = text.strip().replace(",", "").replace("%", "").replace(" ", "").replace("\n", "")
    if cleaned in ("-", "–", "—", "", "N/A", "해당없음"):
        return None
    try:
        val = float(cleaned)
        if 0 <= val <= 100:
            return val
        return None
    except (ValueError, TypeError):
        return None


# ============================================================
# 공개 API 함수들
# ============================================================

def create_delinquency_excel(
    download_path: str,
    output_path: Optional[str] = None,
    api_key: str = None,
    log_callback=None
) -> Optional[str]:
    """
    다운로드 폴더의 통일경영공시 PDF들에서 연체율을 추출하여 엑셀 파일을 생성한다.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    pdf_files = _find_disclosure_pdfs(download_path)
    if not pdf_files:
        log("연체율 추출 대상 통일경영공시 PDF 파일이 없습니다.")
        return None

    method = "Gemini OCR" if (api_key and GEMINI_AVAILABLE) else ("pdfplumber" if PDFPLUMBER_AVAILABLE else None)
    if not method:
        log("pdfplumber와 Gemini 모두 사용 불가하여 연체율 추출을 건너뜁니다.")
        return None

    log(f"연체율 추출 시작: {len(pdf_files)}개 PDF ({method})")
    t0 = time.time()

    # 병렬 추출 (Gemini 사용 시 최대 5개 동시, pdfplumber는 3개)
    max_workers = 5 if (api_key and GEMINI_AVAILABLE) else 3
    results_map = {}  # bank_name → data

    def _extract_one(bank_name, pdf_path):
        return bank_name, extract_delinquency_from_pdf(pdf_path, api_key=api_key, log_callback=log_callback)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_extract_one, bn, pp): bn for bn, pp in pdf_files}
        for future in as_completed(futures):
            try:
                bank_name, data = future.result()
                results_map[bank_name] = data
            except Exception as e:
                bn = futures[future]
                results_map[bn] = None
                log(f"  ❌ {bn}: 추출 오류 - {e}")

    # 결과를 원래 순서대로 정렬
    rows = []
    for bank_name, _ in pdf_files:
        data = results_map.get(bank_name)
        if data:
            rows.append({
                "No": len(rows) + 1,
                "은행명": bank_name,
                "연체율_전기(%)": data.get("연체율_전기", ""),
                "연체율_당기(%)": data.get("연체율_당기", ""),
            })
        else:
            rows.append({
                "No": len(rows) + 1,
                "은행명": bank_name,
                "연체율_전기(%)": "",
                "연체율_당기(%)": "",
            })

    elapsed = time.time() - t0
    elapsed_str = f"{int(elapsed // 60)}분 {int(elapsed % 60)}초" if elapsed >= 60 else f"{elapsed:.1f}초"

    if not rows:
        log("추출된 연체율 데이터가 없습니다.")
        return None

    df = pd.DataFrame(rows)

    if not output_path:
        output_path = os.path.join(
            download_path,
            f"연체율_요약_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="연체율", index=False)
            ws = writer.sheets["연체율"]
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 16

        extracted_count = sum(1 for r in rows if r["연체율_당기(%)"])
        log(f"연체율 엑셀 생성 완료: {extracted_count}/{len(rows)}개 은행 ({elapsed_str})")
        return output_path

    except Exception as e:
        log(f"연체율 엑셀 생성 오류: {e}")
        return None


def extract_all_delinquency(
    download_path: str,
    api_key: str = None,
    log_callback=None
) -> Dict[str, Dict[str, str]]:
    """
    다운로드 폴더의 모든 통일경영공시 PDF에서 연체율을 추출한다.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    pdf_files = _find_disclosure_pdfs(download_path)
    if not pdf_files:
        log("연체율 추출 대상 통일경영공시 PDF 파일이 없습니다.")
        return {}

    method = "Gemini OCR" if (api_key and GEMINI_AVAILABLE) else ("pdfplumber" if PDFPLUMBER_AVAILABLE else None)
    if not method:
        log("pdfplumber와 Gemini 모두 사용 불가합니다.")
        return {}

    log(f"연체율 추출 시작: {len(pdf_files)}개 PDF ({method})")
    t0 = time.time()

    # 병렬 추출
    max_workers = 5 if (api_key and GEMINI_AVAILABLE) else 3
    result = {}

    def _extract_one(bank_name, pdf_path):
        return bank_name, extract_delinquency_from_pdf(pdf_path, api_key=api_key, log_callback=log_callback)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_extract_one, bn, pp): bn for bn, pp in pdf_files}
        for future in as_completed(futures):
            try:
                bank_name, data = future.result()
                if data:
                    result[bank_name] = data
            except Exception as e:
                bn = futures[future]
                log(f"  ❌ {bn}: 추출 오류 - {e}")

    elapsed = time.time() - t0
    elapsed_str = f"{int(elapsed // 60)}분 {int(elapsed % 60)}초" if elapsed >= 60 else f"{elapsed:.1f}초"
    log(f"연체율 추출 완료: {len(result)}/{len(pdf_files)}개 성공 ({elapsed_str})")
    return result


def patch_excel_with_delinquency(
    excel_path: str,
    delinquency_data: Dict[str, Dict[str, str]],
    log_callback=None
) -> bool:
    """기존 분기총괄 엑셀의 연체율 컬럼에 PDF에서 추출한 연체율을 기입한다."""
    if not delinquency_data or not excel_path or not os.path.exists(excel_path):
        return False

    def log(msg):
        if log_callback:
            log_callback(msg)

    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        if "분기총괄" not in wb.sheetnames:
            log("분기총괄 시트를 찾을 수 없습니다.")
            wb.close()
            return False

        ws = wb["분기총괄"]

        header_row = [cell.value for cell in ws[1]]
        company_col = None
        prior_col = None
        current_col = None

        for idx, val in enumerate(header_row):
            if not val:
                continue
            val_str = str(val).strip()
            if val_str == "회사명":
                company_col = idx
            elif "연체율" in val_str and any(kw in val_str for kw in ["전년동기", "전기"]):
                prior_col = idx
            elif "연체율" in val_str and any(kw in val_str for kw in ["금분기", "금기", "당기"]):
                current_col = idx

        if company_col is None:
            log("회사명 컬럼을 찾을 수 없습니다.")
            wb.close()
            return False

        if prior_col is None and current_col is None:
            log("연체율 컬럼을 찾을 수 없습니다.")
            wb.close()
            return False

        patched = 0
        for row_idx in range(2, ws.max_row + 1):
            bank_name = ws.cell(row=row_idx, column=company_col + 1).value
            if not bank_name:
                continue
            bank_name = str(bank_name).strip()

            matched_data = None
            if bank_name in delinquency_data:
                matched_data = delinquency_data[bank_name]
            else:
                for pdf_bank, data in delinquency_data.items():
                    if pdf_bank in bank_name or bank_name in pdf_bank:
                        matched_data = data
                        break
                    clean_pdf = pdf_bank.replace("저축은행", "").strip()
                    clean_bank = bank_name.replace("저축은행", "").strip()
                    if clean_pdf and clean_bank and (clean_pdf in clean_bank or clean_bank in clean_pdf):
                        matched_data = data
                        break

            if not matched_data:
                continue

            updated = False
            if prior_col is not None and matched_data.get("연체율_전기"):
                cell = ws.cell(row=row_idx, column=prior_col + 1)
                existing = cell.value
                if not existing or str(existing).strip() in ("", "-", "0", "0.0"):
                    try:
                        cell.value = float(matched_data["연체율_전기"])
                    except (ValueError, TypeError):
                        cell.value = matched_data["연체율_전기"]
                    updated = True

            if current_col is not None and matched_data.get("연체율_당기"):
                cell = ws.cell(row=row_idx, column=current_col + 1)
                existing = cell.value
                if not existing or str(existing).strip() in ("", "-", "0", "0.0"):
                    try:
                        cell.value = float(matched_data["연체율_당기"])
                    except (ValueError, TypeError):
                        cell.value = matched_data["연체율_당기"]
                    updated = True

            if updated:
                patched += 1

        wb.save(excel_path)
        wb.close()

        log(f"  연체율 기입 완료: {patched}개 은행 반영")
        return patched > 0

    except Exception as e:
        log(f"연체율 엑셀 기입 오류: {e}")
        return False


def _find_disclosure_pdfs(download_path: str) -> List[Tuple[str, str]]:
    """다운로드 폴더에서 통일경영공시 PDF 파일을 찾아 (은행명, 경로) 리스트를 반환한다."""
    results = []
    if not os.path.isdir(download_path):
        return results

    for filename in sorted(os.listdir(download_path)):
        if "통일경영공시" not in filename:
            continue
        if not filename.lower().endswith(".pdf"):
            continue

        filepath = os.path.join(download_path, filename)
        if not os.path.isfile(filepath):
            continue

        bank_name = filename.split("_통일경영공시")[0]
        if bank_name:
            results.append((bank_name, filepath))

    return results
